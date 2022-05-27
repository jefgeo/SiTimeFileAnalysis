import pandas as pd
import os
from openpyxl import load_workbook

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', 100)

writer = pd.ExcelWriter("summary.xlsx")

directory = './Data/'

summary_columns = ['File_Name', 'Column_Name', 'Inferred_Type',
                   'Populated', 'Unique_Values', 'PctUnique',
                   'Nulls', 'PctMissing',
                   'Top_Values', 'mean', 'median', 'min', 'max']
file_count = 0
sheet_count = 0
for file_name in os.listdir(directory):

    summary = []

    if file_name.lower().endswith(".csv") \
            or file_name.lower().endswith(".xlsx") \
            or file_name.lower().endswith(".xls"):
        print(file_name)
        file_count += 1
        if file_name.lower().endswith(".xlsx"):
            xls = load_workbook('{}{}'.format(directory, file_name), read_only=True, keep_links=False)
            sheet_list = xls.sheetnames
        else:
            sheet_list = ['__Single_Sheet__']

        for sheet in sheet_list:
            sheet_count += 1
            print(sheet)
            if file_name.lower().endswith(".csv"):
                df = pd.read_csv('{}{}'.format(directory, file_name), low_memory=False, quotechar='"')
            elif file_name.lower().endswith(".xlsx"):
                df = pd.read_excel('{}{}'.format(directory, file_name), sheet_name=sheet)
            elif file_name.lower().endswith(".xls"):
                df = pd.read_excel('{}{}'.format(directory, file_name))

            top_freq = {}

            for col in df:
                col_name = df[col].name
                col_name_lower = col_name.lower()
                col_counts = df[col].value_counts().to_dict()

                stats = {}
                if df[col].dtypes in ["float64", "int64", "datetime64[ns]"]:
                    stats["mean"] = (df[col].mean())
                    stats["min"] = (df[col].min())
                    stats["max"] = (df[col].max())
                    stats["median"] = (df[col].median())

                unique_values = len(col_counts)

                missing = df[col].isnull()
                missing_count = 0
                for v in missing:
                    if v:
                        missing_count += 1

                {k: v for k, v in sorted(col_counts.items(), key=lambda item: item[1])}
                top_values = list(col_counts)[:unique_values if unique_values < 10 else 5]

                entry = [file_name + ' | ' + sheet if sheet != '__Single_Sheet__' else file_name,
                         col_name_lower, df[col].dtypes,
                         len(df) - missing_count,
                         unique_values, unique_values / len(df) if len(df) > 0 else 0,
                         missing_count, missing_count / len(df) if len(df) > 0 else 0,
                         top_values,
                         (stats["mean"] if "mean" in stats else ''),
                         (stats["median"] if "median" in stats else ''),
                         (stats["min"] if "min" in stats else ''),
                         (stats["max"] if "max" in stats else '')]
                summary.append(entry)

        summary_df = pd.DataFrame(summary, columns=summary_columns)
        summary_df.to_excel(writer, header=True, sheet_name=file_name[0:30], index=False)

        percentage_format = writer.book.add_format({'num_format': '0%'})
        int_format = writer.book.add_format({'num_format': '#,##0'})
        float_format = writer.book.add_format({'num_format': '#,##0.00'})

        for column in summary_df:
            column_width = max(summary_df[column].astype(str).map(len).max(), len(column))
            col_idx = summary_df.columns.get_loc(column)
            writer.sheets[file_name[0:30]].set_column(col_idx, col_idx, column_width)
            if "pct" in column.lower():
                writer.sheets[file_name[0:30]].set_column(col_idx, col_idx, column_width, percentage_format)
            elif any(x in column.lower() for x in ['mean', 'median', 'max', 'min']):
                writer.sheets[file_name[0:30]].set_column(col_idx, col_idx, column_width + 3, float_format)
            elif summary_df[column].dtypes == "int64":
                writer.sheets[file_name[0:30]].set_column(col_idx, col_idx, column_width, int_format)

print('Processed {} Files with a total of {} Sheets'.format(file_count, sheet_count))
writer.close()
